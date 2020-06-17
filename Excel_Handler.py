from openpyxl import load_workbook
from datetime import datetime

from kivy.uix.popup import Popup
from kivy.uix.label import Label

def get_list_size(sheet):
    temp = 2
    while str(sheet.cell(row=temp, column=1).value) != "flag_letzte_Zeile":
        temp += 1
    return temp

def create_sub_lists():
    lst = ['bier', 'nalk', 'wein', 'other']
    lst_size = get_list_size(Pyxl.ws)
    lstp2 = []
    i = 0
    x = 2
    while x < lst_size:
        lst_buff = []
        while Pyxl.ws.cell(row=x, column=2).value == lst[i]:
            lst_buff.append(Pyxl.ws.cell(row=x, column=1).value)
            x += 1
        i += 1
        lstp2.append(lst_buff)
    return lstp2


class Global:
    total = 0  # Aktuelle Summe
    text = ''  # Output-Text des Warenkorb

    product_idx = {}  # Buffer für die Produktliste
    products = {}  # Produktliste mit normalen Preisen
    emp_products = {}  # Produktliste mit Mitarbeiterpreisen

    commission = False  # Boolean, der angibt, ob es sich um Kommissionskauf handelt

    debtor_id = ''  # Name des Kommissionskäufers
    debts = 0  # Summe die insgesamt von Person A auf Kommission gekauft wurde
    debt_payed = False  # Boolean, der angibt, ob die Schulden beglichen wurden

    le_total = '' # Letzter Eintrag zu aktueller Summe
    le_text = '' # Letzer Eintrag zu Output-Text


class Pyxl:
    wb = load_workbook(filename='Bierkeller.xlsx', data_only=True)
    sheets = wb.sheetnames      # Liste aller Sheets in Excel
    ws = wb[sheets[0]]          # Normale Preisliste
    emp = wb[sheets[1]]         # Preisliste für Mitarbeiter
    unused = wb[sheets[2]]      # ungenutzt, zB Preisliste für Externe
    log = wb[sheets[3]]         # Log_Sheet wo Umsatz und Käufe auf Rechnung eingetragen werden
    stat = wb[sheets[4]]        # ungenutzt, Sheet für Statistiken
    active_excel = ws           # Das aktuell aktive Sheet in Workbook

    # Eintrag über Bestellungssumme und Zeitpunkt in Excel-Log eintragen
    def logEntry(self, entry):
        lst_size = get_list_size(Pyxl.log)
        dt_str = datetime.now().strftime("%d.%m.%Y %H:%M")

        if Global.commission is True:
            if entry <= 0:
                Popup(title='ERROR', title_size=20, title_align='center',
                      content=Label(text='Betrag muss positiv sein!',
                                    ont_size=20, halign='center'), size_hint=(.3, .2)).open()
                return
            Pyxl.log.cell(row=lst_size, column=4).value = Global.text
            Pyxl.log.cell(row=lst_size, column=3).value = Global.debtor_id
        elif entry == 0 and len(Global.text) <= 13:
            Popup(title='ERROR', title_size=20, title_align='center',
                  content=Label(text='Keine Eingabe getätigt!',
                                font_size=20, halign='center'), size_hint=(.3, .2)).open()
            return
        Pyxl.log.cell(row=(lst_size + 1), column=1).value = "flag_letzte_Zeile"
        Pyxl.log.cell(row=lst_size, column=2).value = entry
        Pyxl.log.cell(row=lst_size, column=1).value = dt_str
        Pyxl.wb.save(filename='Bierkeller.xlsx')

    def log_payed_account(self, name, total):
        lst_size = get_list_size(Pyxl.log)
        if Global.debt_payed is True:
            for i in range(2, lst_size):
                if Pyxl.log.cell(row=i, column=3).value == name:
                    Pyxl.log.cell(row=i, column=3).value = 'BEZAHLT'
                    Pyxl.wb.save(filename='Bierkeller.xlsx')

    # Erstellt Dictonary basierend auf der Preisliste
    # Rückgabewert ist ein Tupel, welches zwei Dicts enthält, die gleich sind, abgesehen davon, dass key und value
    # vertauscht sind. Bsp. Global.product_idx[0][2] = 'Augustiner', Global.product_idx[1]['Augustiner'] = 2
    def get_product_nr(self):
        lst = []
        lst_size = get_list_size(Pyxl.ws)
        for i in range(2, lst_size):
            lst.append(Pyxl.ws.cell(row=i, column=1).value)
        lst_bf_1 = {j: lst[j - 2] for j in range(2, len(lst) + 2)}
        lst_bf_2 = {lst[j - 2]: j for j in range(2, len(lst) + 2)}
        Global.products = (lst_bf_1, lst_bf_2)
        Pyxl.get_emp_product_nr(Pyxl, lst_size)
        return Global.products

    def get_emp_product_nr(self, lst_size):
        lst = []
        for i in range(2, lst_size):
            lst.append(Pyxl.emp.cell(row=i, column=1).value)
        lst_bf_1 = {j: lst[j - 2] for j in range(2, len(lst) + 2)}
        lst_bf_2 = {lst[j - 2]: j for j in range(2, len(lst) + 2)}
        Global.emp_products = (lst_bf_1, lst_bf_2)
        return Global.emp_products

    def format_text(self, number, type, _str, char):
        if _str is 0:
            return str(char) + '  ' + str(number) + 'x ' + type + '-Flaschen\n'
        elif _str is 1:
            return str(char) + '  ' + str(number) + 'x ' + type + '-Kästen\n'
        else:
            return str(char) + '  ' + str(number) + 'x ' + type + '\n'
